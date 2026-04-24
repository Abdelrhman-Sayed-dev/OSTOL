"""
import_excel.py — استيراد بيانات المركبات والسائقين من Excel إلى قاعدة البيانات
=====================================================================================
الاستخدام:
    python import_excel.py                          # يقرأ كل ملفات .xlsx في نفس المجلد
    python import_excel.py file1.xlsx file2.xlsx    # ملفات محددة
    python import_excel.py --dry-run                # معاينة فقط بدون حفظ

المتطلبات:
    pip install openpyxl

ملاحظات:
    - كلمة المرور الافتراضية لكل سائق جديد: Driver@1234
    - يتجاهل المركبات/السائقين الموجودين مسبقاً
    - يربط كل سائق بمركبته تلقائياً (صلاحية قيادة)
=====================================================================================
"""

import os
import sys
import sqlite3
import bcrypt
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ مطلوب: pip install openpyxl")
    sys.exit(1)


# ── تكوين ──────────────────────────────────────────────────
DATABASE_PATH   = os.environ.get("DATABASE_PATH", "trip_tracker.db")
DEFAULT_PASSWORD = "Driver@1234"
DRY_RUN         = "--dry-run" in sys.argv


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt(rounds=10)).decode()


def fmt_date(val) -> str:
    """تحويل أي قيمة تاريخ إلى YYYY-MM-DD."""
    if not val:
        return ""
    if hasattr(val, 'strftime'):          # datetime object
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    # يعالج: 2026/9/4  أو  2026-9-4  أو  4/9/2026
    for sep in ['/', '-', '.']:
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                if len(parts[0]) == 4:    # YYYY/MM/DD
                    return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
                else:                      # DD/MM/YYYY
                    return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
    return s


def clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


# ══════════════════════════════════════════════════════════════
# استخراج بيانات من ورقة Excel
# ══════════════════════════════════════════════════════════════

PLATE_COLS  = ['رقم المرور', 'رقم اللوحة', 'لوحة']
MODEL_COLS  = ['الماركة', 'الماركة ', 'ماركة', 'النوع']
TYPE_COLS   = ['نوع السيارة', 'نوع السيارة ']
CODE_COLS   = ['الكود', 'رقم الكود', 'كود']
CHASSIS_COLS= ['شاسيه', 'الشاسية', 'رقم الشاسية']
ENGINE_COLS = ['محرك', 'رقم المحرك']
YEAR_COLS   = ['سنة الصنع']
PROJECT_COLS= ['المشروع']
BRANCH_COLS = ['الإدارة / الفرع', 'الفرع / الادارة', 'الإدارة', 'الفرع']
CAR_EXP_COLS= ['انتهاء رخصة السيارة', 'تاريخ انتهاء رخصة السيارة', 'انتهاء رخصة']
DRV_NAME_COLS=['اسم السائق', 'أسم السائق', 'السائق']
DRV_PHONE_COLS=['تليفون السائق', 'رقم التليفون', 'التليفون']
DRV_FIXED_COLS=['رقم ثابت', 'الرقم الثابت']
DRV_EXP_COLS= ['انتهاء رخصة السائق', 'تاريخ انتهاء رخصة السائق']


def pick(row: dict, *col_lists) -> str:
    """يختار أول عمود موجود من القوائم المعطاة."""
    for cols in col_lists:
        for col in cols:
            val = row.get(col, '')
            if val and str(val).strip() not in ('', 'None'):
                return clean(val)
    return ''


def extract_records_from_sheet(ws) -> list[dict]:
    """يستخرج بيانات من ورقة ببنية مرنة."""
    records = []

    # البحث عن صف الرأس (أول صف يحتوي على رقم المرور أو اسم السائق)
    header_row_idx = None
    all_rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(all_rows):
        row_vals = [str(v) if v else '' for v in row]
        for marker in PLATE_COLS + DRV_NAME_COLS:
            if marker in row_vals:
                header_row_idx = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        return []

    # بناء الـ headers map
    headers = [str(v).strip() if v else '' for v in all_rows[header_row_idx]]

    # معالجة صفوف البيانات
    for row_vals in all_rows[header_row_idx + 1:]:
        row = {headers[i]: row_vals[i] for i in range(min(len(headers), len(row_vals)))}
        plate = pick(row, PLATE_COLS)
        if not plate or plate in ('م', '#', 'None'):
            continue

        # تجاهل صفوف المجموع والفراغات
        if any(kw in str(plate) for kw in ['إجمالي', 'مجموع', '—', '---']):
            continue

        records.append({
            'plate':                plate,
            'model':                pick(row, MODEL_COLS) or 'غير محدد',
            'car_name':             pick(row, TYPE_COLS),
            'car_code':             pick(row, CODE_COLS),
            'chassis':              pick(row, CHASSIS_COLS),
            'engine_number':        pick(row, ENGINE_COLS),
            'year':                 pick(row, YEAR_COLS),
            'project':              pick(row, PROJECT_COLS),
            'branch':               pick(row, BRANCH_COLS),
            'car_license_expiry':   fmt_date(pick(row, CAR_EXP_COLS)),
            'driver_name':          pick(row, DRV_NAME_COLS),
            'driver_phone':         pick(row, DRV_PHONE_COLS),
            'driver_fixed_number':  pick(row, DRV_FIXED_COLS),
            'driver_license_expiry': fmt_date(pick(row, DRV_EXP_COLS)),
        })

    return records


# ══════════════════════════════════════════════════════════════
# قراءة كل ملفات Excel
# ══════════════════════════════════════════════════════════════

def read_excel_files(paths: list[str]) -> list[dict]:
    all_records = []
    for path in paths:
        print(f"\n📂 قراءة: {Path(path).name}")
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for sh in wb.sheetnames:
                ws = wb[sh]
                if ws.max_row < 3:
                    continue
                recs = extract_records_from_sheet(ws)
                print(f"   📄 ورقة '{sh}': {len(recs)} سجل")
                all_records.extend(recs)
        except Exception as e:
            print(f"   ❌ خطأ: {e}")

    # إزالة المكرر بناءً على رقم اللوحة
    seen   = set()
    unique = []
    for r in all_records:
        if r['plate'] not in seen:
            seen.add(r['plate'])
            unique.append(r)
        else:
            print(f"   ⚠️  لوحة مكررة تم تجاهلها: {r['plate']}")

    return unique


# ══════════════════════════════════════════════════════════════
# الاستيراد إلى قاعدة البيانات
# ══════════════════════════════════════════════════════════════

def import_to_db(records: list[dict]):
    if not os.path.exists(DATABASE_PATH):
        print(f"❌ قاعدة البيانات غير موجودة: {DATABASE_PATH}")
        sys.exit(1)

    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA journal_mode=WAL")
    c = conn.cursor()

    stats = {
        'cars_added':    0, 'cars_skipped': 0,
        'drivers_added': 0, 'drivers_skipped': 0,
        'perms_added':   0, 'errors': []
    }

    try:
        for rec in records:
            car_id = None

            # ── 1. إدراج المركبة ──────────────────────────────
            try:
                c.execute("""INSERT OR IGNORE INTO cars
                    (plate, model, car_name, car_code, chassis, engine_number,
                     year, project, branch, car_license_expiry, status)
                    VALUES (?,?,?,?,?,?,?,?,?,?,'available')""",
                    (rec['plate'], rec['model'], rec['car_name'],
                     rec['car_code'], rec['chassis'], rec['engine_number'],
                     rec['year'], rec['project'], rec['branch'],
                     rec['car_license_expiry']))

                if c.rowcount > 0:
                    car_id = c.lastrowid
                    stats['cars_added'] += 1
                    print(f"   🚗 + مركبة: {rec['plate']} ({rec['model']})")
                else:
                    c.execute("SELECT id FROM cars WHERE plate=?", (rec['plate'],))
                    row = c.fetchone()
                    car_id = row['id'] if row else None
                    stats['cars_skipped'] += 1
                    print(f"   ⏭  مركبة موجودة: {rec['plate']}")

            except Exception as e:
                stats['errors'].append(f"مركبة {rec['plate']}: {e}")
                continue

            # ── 2. إدراج السائق إن وُجد ──────────────────────
            if not rec['driver_name']:
                continue

            drv_name  = rec['driver_name']
            drv_phone = rec['driver_phone']
            fixed_no  = rec['driver_fixed_number']
            drv_exp   = rec['driver_license_expiry']

            # اسم المستخدم: الرقم الثابت أو الهاتف أو الاسم_مع_الأندرسكور
            username = (
                fixed_no.strip()  if fixed_no.strip()  else
                drv_phone.strip() if drv_phone.strip()  else
                drv_name.replace(' ', '_')
            )

            try:
                # تحقق من وجود المستخدم
                c.execute("SELECT id FROM users WHERE username=?", (username,))
                existing = c.fetchone()

                if existing:
                    user_id = existing['id']
                    stats['drivers_skipped'] += 1
                    print(f"   ⏭  مستخدم موجود: {username}")
                else:
                    # إنشاء المستخدم
                    c.execute("INSERT INTO users(username,password,role) VALUES(?,?,?)",
                              (username, hash_password(DEFAULT_PASSWORD), 'driver'))
                    user_id = c.lastrowid

                    # إنشاء السائق
                    c.execute("""INSERT INTO drivers
                        (name, phone, status, user_id, fixed_number,
                         driver_license_expiry, branch)
                        VALUES (?,?,?,?,?,?,?)""",
                        (drv_name, drv_phone, 'active', user_id,
                         fixed_no, drv_exp, rec['branch']))
                    stats['drivers_added'] += 1
                    print(f"   👤 + سائق: {drv_name} | يوزر: {username}")

                # ── 3. صلاحية قيادة المركبة ──────────────────
                if car_id:
                    c.execute("SELECT id FROM drivers WHERE user_id=?", (user_id,))
                    drv_row = c.fetchone()
                    if drv_row:
                        c.execute("""INSERT OR IGNORE INTO driver_car_permissions
                                     (driver_id, car_id) VALUES (?,?)""",
                                  (drv_row['id'], car_id))
                        if c.rowcount > 0:
                            stats['perms_added'] += 1

            except Exception as e:
                stats['errors'].append(f"سائق {drv_name}: {e}")

        # ── سجل في audit_logs ──────────────────────────────────
        if not DRY_RUN:
            now = datetime.utcnow().isoformat() + "Z"
            c.execute("""INSERT INTO audit_logs(user_id,username,role,action,details,created_at)
                         VALUES(0,'import_script','system','bulk_import',?,?)""",
                      (f"استيراد Excel: {stats['cars_added']} مركبة | {stats['drivers_added']} سائق | {stats['perms_added']} صلاحية",
                       now))
            conn.commit()
            print("\n✅ تم الحفظ في قاعدة البيانات")
        else:
            conn.rollback()
            print("\n🔍 DRY RUN — لم يتم حفظ أي شيء")

    except Exception as e:
        conn.rollback()
        print(f"\n❌ خطأ عام: {e}")
    finally:
        conn.close()

    return stats


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  📦 استيراد بيانات Excel → Fleet Management DB")
    print("=" * 60)

    if DRY_RUN:
        print("  ⚠️  وضع المعاينة فقط — لن يتم حفظ أي بيانات")
        print("=" * 60)

    # تحديد الملفات
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    if args:
        paths = args
    else:
        # جمع كل xlsx في نفس مجلد السكريبت
        base = Path(__file__).parent
        paths = [str(p) for p in base.glob("*.xlsx")]
        if not paths:
            print("❌ لا توجد ملفات .xlsx في نفس المجلد")
            print("   استخدم: python import_excel.py ملف.xlsx")
            sys.exit(1)

    print(f"\n📋 الملفات: {len(paths)}")
    for p in paths:
        print(f"   • {p}")

    # قراءة البيانات
    records = read_excel_files(paths)
    print(f"\n📊 إجمالي السجلات المستخرجة: {len(records)}")
    print(f"   منها سائقين: {len([r for r in records if r['driver_name']])}")

    if not records:
        print("❌ لا توجد بيانات للاستيراد")
        sys.exit(0)

    # معاينة أولى 5 سجلات
    print("\n🔍 عينة من البيانات:")
    for r in records[:5]:
        print(f"   🚗 {r['plate']:20} | {r['model']:15} | 👤 {r['driver_name'] or '—'}")

    # تأكيد (إلا في dry-run)
    if not DRY_RUN:
        confirm = input(f"\n❓ هل تريد استيراد {len(records)} سجل؟ (نعم/لا): ").strip()
        if confirm not in ('نعم', 'y', 'yes', 'Y'):
            print("تم الإلغاء.")
            sys.exit(0)

    # الاستيراد
    stats = import_to_db(records)

    # النتائج
    print("\n" + "=" * 60)
    print("  📊 نتائج الاستيراد")
    print("=" * 60)
    print(f"  🚗 مركبات أضيفت:        {stats['cars_added']}")
    print(f"  ⏭  مركبات موجودة سابقاً: {stats['cars_skipped']}")
    print(f"  👤 سائقين أضيفوا:        {stats['drivers_added']}")
    print(f"  ⏭  سائقين موجودين:       {stats['drivers_skipped']}")
    print(f"  🔓 صلاحيات أضيفت:        {stats['perms_added']}")
    if stats['errors']:
        print(f"  ⚠️  أخطاء ({len(stats['errors'])}):")
        for e in stats['errors'][:10]:
            print(f"     - {e}")
    print("=" * 60)

    if not DRY_RUN and stats['drivers_added'] > 0:
        print(f"\n  🔑 كلمة المرور الافتراضية لكل سائق جديد: {DEFAULT_PASSWORD}")
        print("  ⚠️  يُنصح بتغييرها فور تسليم الأجهزة!")


if __name__ == "__main__":
    main()

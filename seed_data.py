"""seed_data.py — wipes old data, imports drivers & cars from Excel files"""
import sqlite3, bcrypt, os, re
import openpyxl
from datetime import datetime

DB = os.environ.get("DATABASE_PATH", "trip_tracker.db")

# ══ مسارات ملفات الإكسل ══
EXCEL_FILES = {
    "المنشآت المتميزة": "إدراج_الشاسية_رخصة_السير_ورخصة_القيادة_المنشات_المتميزة_شهر4-2026.xlsx",
    "فرع القاهرة":      "بيان_بسيارات_الركوب_التابعة_لفرع_القاهرة_١٠٥٣٣٧.xlsx",
    "صيانة القصور":     "بيان_تراخيص_سيارات_وسائق.xlsx",
}


def hp(pw):
    return bcrypt.hashpw(str(pw).encode(), bcrypt.gensalt()).decode()


def fix_date(val):
    """يحول أي صيغة تاريخ إلى YYYY-MM-DD"""
    if val is None:
        return ''
    if isinstance(val, datetime):
        if val.year > 2100 or val.year < 1900:
            return ''
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    for pat, fmt in [
        (r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$', 'ymd'),
        (r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$', 'dmy'),
    ]:
        m = re.match(pat, s)
        if m:
            a, b, c = int(m.group(1)), int(m.group(2)), int(m.group(3))
            y, mo, d = (a, b, c) if fmt == 'ymd' else (c, b, a)
            if y > 2100 or y < 1900:
                return ''
            try:
                return datetime(y, mo, d).strftime('%Y-%m-%d')
            except ValueError:
                return ''
    return ''


def fix_str(v):
    return '' if v is None else str(v).strip()


def fix_num(v):
    return '' if v is None else str(v).strip().split('.')[0]


def fix_phone(v):
    if v is None:
        return ''
    s = str(v).strip().split('.')[0]
    return s if s.startswith('0') else '0' + s


# ══ قراءة الإكسل وبناء DATA list بنفس شكل الأصل ══
DATA = []

def read_file1(path, sector):
    """المنشآت المتميزة:
       col2=نوع السيارة  col3=الكود  col4=الشاسية  col5=الماركة
       col7=رقم المرور  col8=رخصة السيارة
       col9=اسم السائق  col10=الرقم الثابت  col11=تليفون  col12=رخصة السائق"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        if not isinstance(row[0], (int, float)):
            continue
        plate = fix_str(row[7])
        if not plate or plate.startswith('.'):
            continue
        brand    = fix_str(row[5])
        car_type = fix_str(row[2])
        model    = f"{car_type} ( {brand} )" if brand else car_type
        drv_name = fix_str(row[9])
        if drv_name in ('عمرة محرك', 'محرك إستيراد', ''):
            drv_name = None
        fixed_num = fix_num(row[10])
        DATA.append({
            'sector':       sector,
            'car_type':     model,
            'plate':        plate,
            'chassis':      fix_num(row[4]),
            'driver_name':  drv_name,
            'fixed_num':    fixed_num,
            'phone':        fix_phone(row[11]),
            'drv_lic_exp':  fix_date(row[12]),
            'car_lic_exp':  fix_date(row[8]),
            'username':     drv_name.split()[0] if drv_name else None,
            'password':     fixed_num if fixed_num else None,
        })

def read_file2(path, sector):
    """فرع القاهرة:
       col1=نوع السيارة  col2=الماركة  col3=الكود  col4=اللوحة  col5=الشاسية
       col6=اسم السائق  col7=الرقم الثابت  col8=تليفون
       col9=رخصة السيارة  col10=رخصة السائق"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Sheet1']
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        if not any(row[1:6]):
            continue
        plate = fix_str(row[4])
        if not plate:
            continue
        brand    = fix_str(row[2])
        car_type = fix_str(row[1])
        model    = f"{car_type} ( {brand} )" if brand else car_type
        drv_name = fix_str(row[6])
        if drv_name in ('عمرة محرك', ''):
            drv_name = None
        fixed_num = fix_num(row[7])
        DATA.append({
            'sector':       sector,
            'car_type':     model,
            'plate':        plate,
            'chassis':      fix_str(row[5]),
            'driver_name':  drv_name,
            'fixed_num':    fixed_num,
            'phone':        fix_phone(row[8]),
            'drv_lic_exp':  fix_date(row[10]),
            'car_lic_exp':  fix_date(row[9]),
            'username':     drv_name.split()[0] if drv_name else None,
            'password':     fixed_num if fixed_num else None,
        })

def read_file3(path, sector):
    """صيانة القصور:
       col1=الإدارة  col2=نوع السيارة  col3=الكود  col4=الشاسية
       col6=رقم المرور  col7=الماركة
       col9=اسم السائق  col10=الرقم الثابت  col11=تليفون
       col12=رخصة السيارة  col13=رخصة السائق"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        if not isinstance(row[0], (int, float)):
            continue
        plate = fix_str(row[6])
        if not plate:
            continue
        brand    = fix_str(row[7])
        car_type = fix_str(row[2])
        model    = f"{car_type} ( {brand} )" if brand else car_type
        drv_name = fix_str(row[9])
        if drv_name == '':
            drv_name = None
        fixed_num = fix_num(row[10])
        DATA.append({
            'sector':       fix_str(row[1]) or sector,
            'car_type':     model,
            'plate':        plate,
            'chassis':      fix_str(row[4]),
            'driver_name':  drv_name,
            'fixed_num':    fixed_num,
            'phone':        fix_phone(row[11]),
            'drv_lic_exp':  fix_date(row[13]),
            'car_lic_exp':  fix_date(row[12]),
            'username':     drv_name.split()[0] if drv_name else None,
            'password':     fixed_num if fixed_num else None,
        })


# ══ ابحث عن الملفات في نفس مجلد السكريبت أو /mnt/user-data/uploads ══
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SEARCH_DIRS = [SCRIPT_DIR, "/mnt/user-data/uploads"]

def find_file(name):
    for d in SEARCH_DIRS:
        # exact match
        full = os.path.join(d, name)
        if os.path.exists(full):
            return full
        # match with prefix (e.g. timestamp prefix on uploads)
        for f in os.listdir(d):
            if f.endswith(name):
                return os.path.join(d, f)
    raise FileNotFoundError(f"لم يتم العثور على الملف: {name}")


print("📖 قراءة ملفات الإكسل...")
for sector, fname in EXCEL_FILES.items():
    path = find_file(fname)
    print(f"  ✓ {sector}: {os.path.basename(path)}")
    if sector == "المنشآت المتميزة":
        read_file1(path, sector)
    elif sector == "فرع القاهرة":
        read_file2(path, sector)
    elif sector == "صيانة القصور":
        read_file3(path, sector)

print(f"  📋 إجمالي الصفوف: {len(DATA)}")

# ══ فتح قاعدة البيانات ══
conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
conn.execute("PRAGMA foreign_keys=ON")
c = conn.cursor()

# ══ ADD MISSING COLUMNS (safe migration) ══
print("Running migrations...")
migrations = [
    ("cars",    "chassis",              "TEXT DEFAULT ''"),
    ("cars",    "car_license_expiry",   "TEXT DEFAULT ''"),
    ("cars",    "sector",               "TEXT DEFAULT ''"),
    ("drivers", "driver_license_expiry","TEXT DEFAULT ''"),
    ("drivers", "vehicle_license_expiry","TEXT DEFAULT ''"),
    ("drivers", "national_id",          "TEXT DEFAULT ''"),
]
for tbl, col, typ in migrations:
    try:
        c.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} {typ}")
        print(f"  + Added {tbl}.{col}")
    except Exception:
        pass
conn.commit()
print("✓ Migrations done")

# ══ WIPE ALL OLD DATA ══
print("Wiping old data...")
for tbl in ["driver_car_permissions", "workshop_records", "trips",
            "emergency_reports", "garage_records"]:
    c.execute(f"DELETE FROM {tbl}")
c.execute("DELETE FROM drivers")
c.execute("DELETE FROM cars")
c.execute("DELETE FROM users WHERE username NOT IN ("
          "'admin','Eng mohamed mansour','Eng mohamed sayed',"
          "'Eng abdelrhman sayed','admin1','admin2','admin3',"
          "'supre mohamed sayed','super abdelrhman sayed','super mohamed mansour')")
conn.commit()
print("✓ Old data cleared")

# ══ IMPORT ══
# تتبع اليوزرنيمات المستخدمة لتفادي التكرار
used_usernames = {}

for row in DATA:
    # ── Insert car ──
    try:
        c.execute("""INSERT OR IGNORE INTO cars(plate,model,status,chassis,car_license_expiry,sector)
                     VALUES(?,?,?,?,?,?)""",
                  (row["plate"], row["car_type"], "available",
                   row["chassis"], row["car_lic_exp"], row["sector"]))
    except Exception as e:
        print(f"Car {row['plate']}: {e}")

    # ── بدون سائق ──
    if not row["driver_name"]:
        print(f"✓ {row['plate']} | بدون سائق")
        continue

    # ── توليد يوزرنيم فريد ──
    base_uname = row["username"] or row["driver_name"].split()[0]
    if base_uname not in used_usernames:
        used_usernames[base_uname] = 0
        uname = base_uname
    else:
        used_usernames[base_uname] += 1
        uname = base_uname + str(used_usernames[base_uname])

    pw = row["password"] or uname

    # ── Insert user ──
    uid = None
    try:
        c.execute("SELECT id FROM users WHERE username=?", (uname,))
        ex = c.fetchone()
        if not ex:
            c.execute("INSERT INTO users(username,password,role) VALUES(?,?,?)",
                      (uname, hp(pw), "driver"))
            uid = c.lastrowid
        else:
            uid = ex["id"]
    except Exception as e:
        print(f"User {uname}: {e}")
        continue

    # ── Insert driver ──
    did = None
    try:
        c.execute("""INSERT INTO drivers(name,phone,status,user_id,
                     driver_license_expiry,vehicle_license_expiry,national_id)
                     VALUES(?,?,?,?,?,?,?)""",
                  (row["driver_name"], row["phone"], "active", uid,
                   row["drv_lic_exp"], row["car_lic_exp"], row["fixed_num"]))
        did = c.lastrowid
    except Exception as e:
        print(f"Driver {row['driver_name']}: {e}")
        continue

    # ── Link driver ↔ car permission ──
    try:
        c.execute("SELECT id FROM cars WHERE plate=?", (row["plate"],))
        car = c.fetchone()
        if car and did:
            c.execute("INSERT OR IGNORE INTO driver_car_permissions(driver_id,car_id) VALUES(?,?)",
                      (did, car["id"]))
    except Exception as e:
        print(f"Permission: {e}")

    print(f"✓ {row['driver_name']} | {row['plate']} | user:{uname} pw:{pw}")

conn.commit()

# ══ إحصائيات ══
c.execute("SELECT COUNT(*) as n FROM cars")
total_cars = c.fetchone()["n"]
c.execute("SELECT COUNT(*) as n FROM drivers")
total_drivers = c.fetchone()["n"]

conn.close()
print(f"\n✅ Done! {len(DATA)} rows processed — {total_cars} cars, {total_drivers} drivers imported.")
